import re
from openpyxl import Workbook
from openpyxl.chart import LineChart,Reference

# File path
input_file = r'C:\Users\luisa\Desktop\SCS\SCS_Project\outputs.txt'
excel_file = r'C:\Users\luisa\Desktop\SCS\SCS_Project\Graphs.xlsx'

ms_c_execution_times = []
ms_c_th_creation = []
ms_c_context_times = []
ms_c_th_mig_times = []

mm_c_execution_times = []
mm_c_th_creation = []
mm_c_context_times = []
mm_c_th_mig_times = []

ms_py_execution_times = []
ms_py_th_creation = []
ms_py_context_times = []

mm_py_execution_times = []
mm_py_th_creation = []
mm_py_context_times = []

ms_java_execution_times = []
ms_java_th_creation = []
ms_java_context_times = []

mm_java_execution_times = []
mm_java_th_creation = []
mm_java_context_times = []

n_values = []

with open(input_file, 'r') as file:
    for line in file:
        if "(n) n =" in line:
            match = re.search(r'\(n\) n = (\d+)', line)
            if match:
                n_values.append(int(match.group(1)))
        if "(ms-c)" in line:
            match = re.search(r'Average execution time:\s*(\d+\.\d+)', line)
            if match:
                ms_c_execution_times.append(match.group(1))
            match = re.search(r'Average thread creation time:\s*(\d+\.\d+)', line)
            if match:
                ms_c_th_creation.append(match.group(1))
            match = re.search(r'Context switch time:\s*(\d+\.\d+)', line)
            if match:
                ms_c_context_times.append(match.group(1))
            match = re.search(r'Thread migration time:\s*(\d+\.\d+)', line)
            if match:
                ms_c_th_mig_times.append(match.group(1))
        if "(mm-c)" in line:
            match = re.search(r'Average execution time:\s*(\d+\.\d+)', line)
            if match:
                mm_c_execution_times.append(match.group(1))
            match = re.search(r'Average thread creation time:\s*(\d+\.\d+)', line)
            if match:
                mm_c_th_creation.append(match.group(1))
            match = re.search(r'Context switch time:\s*(\d+\.\d+)', line)
            if match:
                mm_c_context_times.append(match.group(1))
            match = re.search(r'Thread migration time:\s*(\d+\.\d+)', line)
            if match:
                mm_c_th_mig_times.append(match.group(1))
        if "(mm-py)" in line:
            match = re.search(r'Execution time:\s*(\d+\.\d+)', line)
            if match:
                mm_py_execution_times.append(match.group(1))
            match = re.search(r'Average thread creation time:\s*(\d+\.\d+)', line)
            if match:
                mm_py_th_creation.append(match.group(1))
            match = re.search(r'Context switch time:\s*(\d+\.\d+)', line)
            if match:
                mm_py_context_times.append(match.group(1))
        if "(ms-py)" in line:
            match = re.search(r'Execution time:\s*(\d+\.\d+)', line)
            if match:
                ms_py_execution_times.append(match.group(1))
            match = re.search(r'Average thread creation time:\s*(\d+\.\d+)', line)
            if match:
                ms_py_th_creation.append(match.group(1))
            match = re.search(r'Context switch time:\s*(\d+\.\d+)', line)
            if match:
                ms_py_context_times.append(match.group(1))
        if "(mm-java)" in line:
            match = re.search(r'Execution time:\s*(\d+\.\d+)', line)
            if match:
                mm_java_execution_times.append(match.group(1))
            match = re.search(r'Average thread creation time:\s*(\d+\.\d+)', line)
            if match:
                mm_java_th_creation.append(match.group(1))
            match = re.search(r'Context switch time:\s*(\d+\.\d+)', line)
            if match:
                mm_java_context_times.append(match.group(1))
        if "(ms-java)" in line:
            match = re.search(r'Execution time:\s*(\d+\.\d+)', line)
            if match:
                ms_java_execution_times.append(match.group(1))
            match = re.search(r'Average thread creation time:\s*(\d+\.\d+)', line)
            if match:
                ms_java_th_creation.append(match.group(1))
            match = re.search(r'Context switch time:\s*(\d+\.\d+)', line)
            if match:
                ms_java_context_times.append(match.group(1))


wb = Workbook()

ws1 = wb.active
ws1.title = "MS_Exec_time"
ws1['A1'] = "C"  # Add index column header
ws1['B1'] = "Python"
ws1['C1'] = "Java"
ws1['D1'] = "n"
for i, execution_time in enumerate(ms_c_execution_times, start=2):
    ws1[f'A{i}'] = float(execution_time)
for i, execution_time in enumerate(ms_py_execution_times, start=2):
    ws1[f'B{i}'] = float(execution_time)
for i, execution_time in enumerate(ms_java_execution_times, start=2):
    ws1[f'C{i}'] = float(execution_time)
for i, n in enumerate(n_values, start=2):
    ws1[f'D{i}'] = int(n)



ws2 = wb.create_sheet(title="MS_Th_creation")
ws2['A1'] = "C"
ws2['B1'] = "Python"
ws2['C1'] = "Java"
ws2['D1'] = "n"
for i, creation_time in enumerate(ms_c_th_creation, start=2):
    ws2[f'A{i}'] = float(creation_time)
for i, creation_time in enumerate(ms_py_th_creation, start=2):
    ws2[f'B{i}'] = float(creation_time)
for i, creation_time in enumerate(ms_java_th_creation, start=2):
    ws2[f'C{i}'] = float(creation_time)
for i, n in enumerate(n_values, start=2):
    ws2[f'D{i}'] = int(n)

ws3 = wb.create_sheet(title="MS_Context")
ws3['A1'] = "C"
ws3['B1'] = "Python"
ws3['C1'] = "Java"
ws3['D1'] = "n"
for i, context_time in enumerate(ms_c_context_times, start=2):
    ws3[f'A{i}'] = float(context_time)
for i, context_time in enumerate(ms_py_context_times, start=2):
    ws3[f'B{i}'] = float(context_time)
for i, context_time in enumerate(ms_java_context_times, start=2):
    ws3[f'C{i}'] = float(context_time)
for i, n in enumerate(n_values, start=2):
    ws3[f'D{i}'] = int(n)

ws4 = wb.create_sheet(title="MS_Th_mig")
ws4['A1'] = "C"
ws4['B1'] = "Python"
ws4['C1'] = "Java"
ws4['D1'] = "n"
for i, mig_time in enumerate(ms_c_th_mig_times, start=2):
    ws4[f'A{i}'] = mig_time

ws5 = wb.create_sheet(title="MM_Exec_time")
ws5['A1'] = "C"
ws5['B1'] = "Python"
ws5['C1'] = "Java"
for i, creation_time in enumerate(mm_c_execution_times, start=2):
    ws5[f'A{i}'] = float(creation_time)
for i, creation_time in enumerate(mm_py_execution_times, start=2):
    ws5[f'B{i}'] = float(creation_time)
for i, creation_time in enumerate(mm_java_execution_times, start=2):
    ws5[f'C{i}'] = float(creation_time)


ws6 = wb.create_sheet(title="MM_Th_creation")
ws6['A1'] = "C"
ws6['B1'] = "Python"
ws6['C1'] = "Java"
for i, creation_time in enumerate(mm_c_th_creation, start=2):
    ws6[f'A{i}'] = float(creation_time)
for i, creation_time in enumerate(mm_py_th_creation, start=2):
    ws6[f'B{i}'] = float(creation_time)
for i, creation_time in enumerate(mm_java_th_creation, start=2):
    ws6[f'C{i}'] = float(creation_time)


ws7 = wb.create_sheet(title="MM_Join")
ws7['A1'] = "C"
ws7['B1'] = "Python"
ws7['C1'] = "Java"
for i, context_time in enumerate(mm_c_context_times, start=2):
    ws7[f'A{i}'] = float(context_time)
for i, context_time in enumerate(mm_py_context_times, start=2):
    ws7[f'B{i}'] = float(context_time)
for i, context_time in enumerate(mm_java_context_times, start=2):
    ws7[f'C{i}'] = float(context_time)


ws8 = wb.create_sheet(title="MM_Th_mig")
ws8['A1'] = "C"
ws8['B1'] = "Python"
ws8['C1'] = "Java"
for i, mig_time in enumerate(mm_c_th_mig_times, start=2):
    ws8[f'A{i}'] = mig_time


def ms_create_chart(ws):
    values = Reference(ws, min_col=1, min_row=1, max_col=ws.max_column - 1, max_row=ws.max_row)
    categories = Reference(ws, min_col=4, min_row=2, max_row=ws.max_row)
    chart = LineChart()
    chart.title = ws.title
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)
    ws.add_chart(chart, "F5")

def mm_create_chart(ws):
    values = Reference(ws, min_col=1, min_row=1, max_col=ws.max_column, max_row=ws.max_row)
    chart = LineChart()
    chart.title = ws.title
    chart.add_data(values, titles_from_data=True)
    ws.add_chart(chart, "F5")

ms_create_chart(ws1)
ms_create_chart(ws2)
ms_create_chart(ws3)
mm_create_chart(ws5)
mm_create_chart(ws6)
mm_create_chart(ws7)

wb.save(excel_file)